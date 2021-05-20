import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from pyexcel.cookbook import merge_all_to_a_book
import glob
import sys


def convertCSV_excel(csvPath):
    temp_path = "temp.xlsx"
    try:
        merge_all_to_a_book(glob.glob(csvPath), temp_path)
        return temp_path
    except Exception as err:
        print("Error opening file: " + str(err))
        sys.exit(0)


def create_excel_obj(temp_excel_path):
    try:
        workbook = openpyxl.load_workbook(temp_excel_path)
        return workbook
    except Exception as err:
        err_output = "Error opening temp excel file: " + str(err)
        display_listbox.delete(0, tk.END)
        display_listbox.insert(0, err_output)

def process_excel(wb):
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            appending_string = ""
            if cell.value == cell.offset(row=1, column=0).value:
                appending_string += str(cell.value)
                i = 1
                while cell.offset(row=i, column=0).value == cell.value and cell.value != None:
                    if str(cell.offset(row=i, column=3).value) not in appending_string:
                        appending_string += str(cell.offset(row=i, column=3).value) + ' '
                    sheet.delete_rows(cell.row + i, 1)
                    i += 1
                cell.offset(row=0, column=3).value = appending_string
    wb.save(filename= 'temp.xlsx')




def delete_temp_excel(temp_file):
    if os.path.isfile(temp_file):
        os.remove(temp_file)



############################## event handlers #######################################

#Event handler for the 'CSV' button
def get_csv():
    filename = filedialog.askopenfilename(title = "Select the CSV to be reformatted")
    csv_textbox.delete(0, tk.END)
    csv_textbox.insert(0, filename)

#Event handler for the 'Close' button
def close_it():
    sys.exit(0)

#Event for 'Run it' button, processes CSV
def run_it():
    csvPath = csv_textbox.get()
    temp_excel = convertCSV_excel(csvPath)
    temp_workbook = create_excel_obj(temp_excel)
    process_excel(temp_workbook)
   # delete_temp_excel(temp_excel)
    







################################ main #########################################

#Create the root window
window = tk.Tk()
window.title("CSV Reformatter - Salesforce to Qualtrics")

#Four frames:
header_frame = tk.Frame(master = window)
csv_script_frame = tk.Frame(master = window)
db_frame = tk.Frame(master = window)
csv_display_frame = tk.Frame(master = window)
run_close_frame = tk.Frame(master = window)

header_frame.pack(side = tk.TOP, fill = tk.BOTH)
csv_script_frame.pack(side = tk.TOP, fill = tk.BOTH)
db_frame.pack(side = tk.TOP, fill = tk.BOTH)
csv_display_frame.pack(side = tk.TOP, fill = tk.BOTH)
run_close_frame.pack(side = tk.TOP)

#Button for picking  CSV
csv_button = tk.Button(master = csv_script_frame, text = "CSV File", command = get_csv)

#Two text boxes for the file paths
csv_textbox = tk.Entry(master = csv_script_frame, width = 100)

csv_button.pack(side = tk.LEFT)
csv_textbox.pack(side = tk.LEFT)

#Listbox for displaying things
display_listbox = tk.Listbox(master = csv_display_frame, width = 113, height = 25)
display_listbox.pack(side = tk.LEFT, fill = tk.Y)

#Scrollbar for the listbox
scrollbar = tk.Scrollbar(master = csv_display_frame)
scrollbar.pack(side = tk.LEFT, fill = tk.Y)

#Associate the scrollbar with the listbox
display_listbox.config(yscrollcommand = scrollbar.set)
scrollbar.config(command = display_listbox.yview) 

#Two buttons, each with its own event_handler
run_button = tk.Button(text = "Run", master = run_close_frame, command = run_it)
close_button = tk.Button(text = "Close", master = run_close_frame, command = close_it)

run_button.pack(side = tk.LEFT)
close_button.pack(side = tk.LEFT)

window.mainloop()  